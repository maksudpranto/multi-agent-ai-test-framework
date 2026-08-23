"""Assemble and serialize the Complete Test Design Package (§10).

One place builds a normalized package dict from a pipeline run's artifacts (the
five primary evaluated outputs: test cases, test data, requirement traceability,
coverage, quality). Serializers render it to JSON / CSV / Markdown / Excel / PDF
so the same content is exported consistently across formats.
"""
from __future__ import annotations

import csv
import io
import json

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    AcceptanceCriterion,
    CoverageReport,
    PipelineRun,
    QualityReport,
    Requirement,
    TestCase,
    TestCaseStatus,
)

FORMATS = {
    "json": ("application/json", "json"),
    "csv": ("text/csv", "csv"),
    "md": ("text/markdown", "md"),
    "xlsx": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "pdf": ("application/pdf", "pdf"),
}


def _current_test_cases(db: Session, run: PipelineRun) -> list[TestCase]:
    all_cases = list(
        db.scalars(
            select(TestCase)
            .where(TestCase.pipeline_run_id == run.id)
            .order_by(TestCase.id)
        )
    )
    superseded = {c.parent_test_case_id for c in all_cases if c.parent_test_case_id}
    current = [
        c
        for c in all_cases
        if c.id not in superseded and c.status != TestCaseStatus.rejected
    ]
    current.sort(key=lambda c: (c.rank if c.rank is not None else 10_000, c.id))
    return current


def build_package(db: Session, run: PipelineRun, requirement: Requirement) -> dict:
    """Normalized package for one run — the shape every serializer consumes."""
    criteria = list(
        db.scalars(
            select(AcceptanceCriterion)
            .where(AcceptanceCriterion.pipeline_run_id == run.id)
            .order_by(AcceptanceCriterion.order)
        )
    )
    cases = _current_test_cases(db, run)
    coverage = list(
        db.scalars(
            select(CoverageReport).where(CoverageReport.pipeline_run_id == run.id)
        )
    )
    quality = {
        q.test_case_id: q
        for q in db.scalars(
            select(QualityReport).where(QualityReport.pipeline_run_id == run.id)
        )
    }
    crit_label = {c.id: f"AC{i + 1}" for i, c in enumerate(criteria)}

    return {
        "requirement": {
            "id": requirement.id,
            "title": requirement.title,
            "type": requirement.req_type.value
            if hasattr(requirement.req_type, "value")
            else requirement.req_type,
            "priority": requirement.priority.value
            if hasattr(requirement.priority, "value")
            else requirement.priority,
            "text": requirement.raw_text,
        },
        "acceptance_criteria": [
            {"id": c.id, "label": crit_label[c.id], "text": c.text} for c in criteria
        ],
        "test_cases": [
            {
                "id": c.id,
                "title": c.title,
                "type": c.type,
                "priority": c.priority,
                "severity": c.severity,
                "rank": c.rank,
                "traces_to": crit_label.get(c.traces_to),
                "steps": c.steps or [],
                "expected_result": c.expected_result,
                "test_data": c.test_data,
                "quality": (
                    {
                        "clarity": quality[c.id].clarity_score,
                        "atomicity": quality[c.id].atomicity_score,
                        "traceability": quality[c.id].traceability_score,
                        "duplicate": quality[c.id].duplicate_flag,
                    }
                    if c.id in quality
                    else None
                ),
            }
            for c in cases
        ],
        "coverage": [
            {
                "criterion": crit_label.get(r.acceptance_criterion_id),
                "covered": r.covered,
                "covering_test_case_ids": r.covering_test_case_ids or [],
                "gap_notes": r.gap_notes,
            }
            for r in coverage
        ],
    }


# --- serializers ----------------------------------------------------------


def _steps_text(steps: list) -> str:
    return " | ".join(f"{i + 1}. {s}" for i, s in enumerate(steps))


def to_json(pkg: dict) -> bytes:
    return json.dumps(pkg, indent=2, ensure_ascii=False).encode("utf-8")


def to_csv(pkg: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(
        [
            "id", "title", "type", "priority", "severity", "rank",
            "traces_to", "steps", "expected_result", "test_data",
            "clarity", "atomicity", "traceability", "duplicate",
        ]
    )
    for tc in pkg["test_cases"]:
        q = tc.get("quality") or {}
        w.writerow(
            [
                tc["id"], tc["title"], tc["type"], tc["priority"],
                tc.get("severity") or "", tc.get("rank") if tc.get("rank") is not None else "",
                tc.get("traces_to") or "", _steps_text(tc["steps"]),
                tc["expected_result"] or "",
                json.dumps(tc.get("test_data")) if tc.get("test_data") else "",
                q.get("clarity", ""), q.get("atomicity", ""),
                q.get("traceability", ""), q.get("duplicate", ""),
            ]
        )
    return buf.getvalue().encode("utf-8")


def _summary(pkg: dict) -> dict:
    """Roll-up numbers shown at the top of every export."""
    cases = pkg["test_cases"]
    cov = pkg["coverage"]
    covered = sum(1 for c in cov if c["covered"])
    total_cov = len(cov)
    qvals: list[float] = []
    for tc in cases:
        q = tc.get("quality")
        if q:
            qvals += [q["clarity"], q["atomicity"], q["traceability"]]
    return {
        "cases": len(cases),
        "criteria": len(pkg["acceptance_criteria"]),
        "covered": covered,
        "total_cov": total_cov,
        "cov_pct": round(100 * covered / total_cov) if total_cov else 0,
        "q_pct": round(100 * sum(qvals) / len(qvals)) if qvals else None,
        "dups": sum(1 for tc in cases if (tc.get("quality") or {}).get("duplicate")),
    }


def _pct(v) -> str:
    return "-" if v is None else f"{round(v * 100)}%"


def _tc_numbers(pkg: dict) -> dict:
    """Map each test case's internal id to its 1-based position, so references
    read as TC1/TC2 (what the reader sees) instead of raw database ids."""
    return {tc["id"]: n for n, tc in enumerate(pkg["test_cases"], 1)}


def to_markdown(pkg: dict) -> bytes:
    r = pkg["requirement"]
    s = _summary(pkg)
    num = _tc_numbers(pkg)
    L = [
        "# Test Design Package",
        f"## {r['title']}",
        "",
        f"_Type: {r['type']} · Generated by MATF (multi-agent test design)_",
        "",
        "### At a glance",
        "",
        "| Metric | Value |",
        "| --- | --- |",
        f"| Test cases | {s['cases']} |",
        f"| Acceptance criteria | {s['criteria']} |",
        f"| Coverage | {s['cov_pct']}% ({s['covered']}/{s['total_cov']} verified) |",
    ]
    if s["q_pct"] is not None:
        L.append(f"| Overall quality | {s['q_pct']}% |")
    if s["dups"]:
        L.append(f"| Possible duplicates | {s['dups']} |")
    L += ["", "## Requirement", "", r["text"], "", "## Acceptance criteria", ""]
    for c in pkg["acceptance_criteria"]:
        L.append(f"- **{c['label']}** — {c['text']}")

    L += ["", "## Test cases", ""]
    for tc in pkg["test_cases"]:
        L.append(f"### TC{num[tc['id']]}. {tc['title']}")
        meta = [f"Type: `{tc['type']}`", f"Priority: {tc['priority']}"]
        if tc.get("severity"):
            meta.append(f"Severity: {tc['severity']}")
        if tc.get("traces_to"):
            meta.append(f"Verifies: {tc['traces_to']}")
        L += [" · ".join(meta), "", "**Steps**", ""]
        for i, step in enumerate(tc["steps"], 1):
            L.append(f"{i}. {step}")
        L += ["", f"**Expected result:** {tc['expected_result']}"]
        if tc.get("test_data"):
            L.append(f"**Test data:** `{json.dumps(tc['test_data'])}`")
        q = tc.get("quality")
        if q:
            ql = (
                f"**Quality:** clarity {_pct(q['clarity'])} · "
                f"atomicity {_pct(q['atomicity'])} · traceability {_pct(q['traceability'])}"
            )
            if q.get("duplicate"):
                ql += " · ⚠ possible duplicate"
            L.append(ql)
        L.append("")

    L += ["## Coverage", "", "| Acceptance criterion | Status | Verified by |", "| --- | --- | --- |"]
    for c in pkg["coverage"]:
        status = "✅ Covered" if c["covered"] else "❌ Gap"
        refs = ", ".join(f"TC{num[i]}" for i in c["covering_test_case_ids"] if i in num)
        verified = refs or (c["gap_notes"] or "—")
        L.append(f"| {c['criterion']} | {status} | {verified} |")
    L.append("")
    return ("\n".join(L) + "\n").encode("utf-8")


def to_xlsx(pkg: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    s = _summary(pkg)
    num = _tc_numbers(pkg)
    r = pkg["requirement"]

    head_fill = PatternFill("solid", fgColor="4F46E5")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

    def widths(ws, *vals) -> None:
        for i, w in enumerate(vals, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # -- Overview sheet --
    ov = wb.active
    ov.title = "Overview"
    ov["A1"] = f"Test Design Package — {r['title']}"
    ov["A1"].font = title_font
    ov["A3"] = "Type"
    ov["B3"] = r["type"]
    rows = [
        ("Test cases", s["cases"]),
        ("Acceptance criteria", s["criteria"]),
        ("Coverage", f"{s['cov_pct']}%  ({s['covered']}/{s['total_cov']} verified)"),
    ]
    if s["q_pct"] is not None:
        rows.append(("Overall quality", f"{s['q_pct']}%"))
    if s["dups"]:
        rows.append(("Possible duplicates", s["dups"]))
    for i, (k, v) in enumerate(rows, start=5):
        ov.cell(row=i, column=1, value=k).font = Font(bold=True)
        ov.cell(row=i, column=2, value=v)
    widths(ov, 22, 46)

    # -- Test Cases sheet --
    ws = wb.create_sheet("Test Cases")
    ws.append(
        ["TC", "Title", "Type", "Priority", "Severity", "Verifies",
         "Steps", "Expected Result", "Test Data",
         "Clarity", "Atomicity", "Traceability", "Duplicate?"]
    )
    for tc in pkg["test_cases"]:
        q = tc.get("quality") or {}
        row = [
            f"TC{num[tc['id']]}", tc["title"], tc["type"], tc["priority"],
            tc.get("severity") or "", tc.get("traces_to") or "",
            "\n".join(f"{i}. {stp}" for i, stp in enumerate(tc["steps"], 1)),
            tc["expected_result"] or "",
            json.dumps(tc.get("test_data")) if tc.get("test_data") else "",
            q.get("clarity"), q.get("atomicity"), q.get("traceability"),
            "Yes" if q.get("duplicate") else "",
        ]
        ws.append(row)
        rr = ws.max_row
        for col in (2, 7, 8, 9):
            ws.cell(row=rr, column=col).alignment = wrap_top
        for col in (10, 11, 12):
            c = ws.cell(row=rr, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = "0%"
    style_header(ws, 13)
    widths(ws, 6, 34, 12, 10, 10, 10, 40, 34, 22, 9, 10, 12, 11)

    # -- Coverage sheet --
    cov = wb.create_sheet("Coverage")
    cov.append(["Acceptance Criterion", "Status", "Verified By", "Gap Notes"])
    for c in pkg["coverage"]:
        refs = ", ".join(f"TC{num[i]}" for i in c["covering_test_case_ids"] if i in num)
        cov.append([c["criterion"], "Covered" if c["covered"] else "GAP", refs, c["gap_notes"] or ""])
        cov.cell(row=cov.max_row, column=1).alignment = wrap_top
        cov.cell(row=cov.max_row, column=4).alignment = wrap_top
    style_header(cov, 4)
    widths(cov, 46, 10, 18, 34)

    # -- Acceptance Criteria sheet --
    crit = wb.create_sheet("Acceptance Criteria")
    crit.append(["Label", "Criterion"])
    for c in pkg["acceptance_criteria"]:
        crit.append([c["label"], c["text"]])
        crit.cell(row=crit.max_row, column=2).alignment = wrap_top
    style_header(crit, 2)
    widths(crit, 10, 80)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_pdf(pkg: dict) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    def clean(s: str) -> str:
        # fpdf core fonts are latin-1 only; drop unsupported chars.
        return (s or "").encode("latin-1", "replace").decode("latin-1")

    r = pkg["requirement"]
    s = _summary(pkg)
    num = _tc_numbers(pkg)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.set_margins(16, 16, 16)
    pdf.add_page()
    ink = (28, 29, 34)
    muted = (108, 112, 122)
    accent = (79, 70, 229)

    def line(text: str, h: float = 6) -> None:
        pdf.multi_cell(0, h, clean(text), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    def heading(text: str) -> None:
        pdf.ln(3)
        pdf.set_text_color(*accent)
        pdf.set_font("Helvetica", "B", 13)
        line(text, 7)
        pdf.set_draw_color(224, 224, 230)
        y = pdf.get_y()
        pdf.line(16, y, 194, y)
        pdf.ln(2)
        pdf.set_text_color(*ink)

    # Title
    pdf.set_text_color(*ink)
    pdf.set_font("Helvetica", "B", 18)
    line(f"Test Design Package", 9)
    pdf.set_font("Helvetica", "", 12)
    line(r["title"], 7)
    pdf.set_text_color(*muted)
    pdf.set_font("Helvetica", "", 9)
    line(f"Type: {r['type']}  |  Generated by MATF (multi-agent test design)")
    pdf.set_text_color(*ink)

    # At a glance
    heading("At a glance")
    pdf.set_font("Helvetica", "", 10)
    line(f"Test cases: {s['cases']}    Acceptance criteria: {s['criteria']}", 5.5)
    line(f"Coverage: {s['cov_pct']}% ({s['covered']}/{s['total_cov']} verified)", 5.5)
    if s["q_pct"] is not None:
        extra = f"    Possible duplicates: {s['dups']}" if s["dups"] else ""
        line(f"Overall quality: {s['q_pct']}%{extra}", 5.5)

    # Requirement
    heading("Requirement")
    pdf.set_font("Helvetica", "", 10)
    line(r["text"], 5.5)

    # Acceptance criteria
    heading("Acceptance criteria")
    pdf.set_font("Helvetica", "", 10)
    for c in pkg["acceptance_criteria"]:
        pdf.set_font("Helvetica", "B", 10)
        pdf.write(5.5, clean(f"{c['label']}  "))
        pdf.set_font("Helvetica", "", 10)
        line(c["text"], 5.5)

    # Test cases
    heading("Test cases")
    for tc in pkg["test_cases"]:
        pdf.set_font("Helvetica", "B", 11)
        line(f"TC{num[tc['id']]}. {tc['title']}", 6)
        pdf.set_text_color(*muted)
        pdf.set_font("Helvetica", "", 9)
        meta = f"{tc['type']}  |  priority {tc['priority']}"
        if tc.get("severity"):
            meta += f"  |  severity {tc['severity']}"
        if tc.get("traces_to"):
            meta += f"  |  verifies {tc['traces_to']}"
        line(meta, 5)
        pdf.set_text_color(*ink)
        pdf.set_font("Helvetica", "", 10)
        for i, stp in enumerate(tc["steps"], 1):
            line(f"   {i}. {stp}", 5.5)
        pdf.set_font("Helvetica", "B", 10)
        pdf.write(5.5, "   Expected: ")
        pdf.set_font("Helvetica", "", 10)
        line(tc["expected_result"], 5.5)
        if tc.get("test_data"):
            pdf.set_text_color(*muted)
            pdf.set_font("Helvetica", "", 9)
            line(f"   Test data: {json.dumps(tc['test_data'])}", 5)
            pdf.set_text_color(*ink)
        q = tc.get("quality")
        if q:
            pdf.set_text_color(*muted)
            pdf.set_font("Helvetica", "", 9)
            ql = (
                f"   Quality: clarity {_pct(q['clarity'])} | atomicity {_pct(q['atomicity'])} "
                f"| traceability {_pct(q['traceability'])}"
            )
            if q.get("duplicate"):
                ql += " | possible duplicate"
            line(ql, 5)
            pdf.set_text_color(*ink)
        pdf.ln(2)

    # Coverage
    heading("Coverage")
    pdf.set_font("Helvetica", "", 10)
    for c in pkg["coverage"]:
        refs = ", ".join(f"TC{num[i]}" for i in c["covering_test_case_ids"] if i in num)
        status = "Covered" if c["covered"] else "GAP"
        tail = refs or (c["gap_notes"] or "-")
        line(f"{c['criterion']}: {status} - {tail}", 5.5)

    return bytes(pdf.output())


SERIALIZERS = {
    "json": to_json,
    "csv": to_csv,
    "md": to_markdown,
    "xlsx": to_xlsx,
    "pdf": to_pdf,
}


def serialize(pkg: dict, fmt: str) -> bytes:
    return SERIALIZERS[fmt](pkg)
